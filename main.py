# - Importação
import tkinter as tk
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# - Funções
def salvar_dados_na_planilha(data_hora, temperatura, umidade):
  # Verifica se o arquivo já existe
  if os.path.exists("historico_temperatura.xlsx"):
    # Abre o arquivo existente
    workbook = load_workbook("historico_temperatura.xlsx")
    sheet = workbook.active
  else:
    # Cria um novo arquivo
    workbook = Workbook()
    sheet = workbook.active
    # Cria cabeçalhos
    sheet.append(["Data e Hora", "Temperatura", "Umidade"])

  # Adiciona a nova linha com os dados
  sheet.append([data_hora, temperatura, umidade])

  # Salva o arquivo
  workbook.save("historico_temperatura.xlsx")

def salvar_temperatura():
  # Inicialização do WebDriver
  driver = webdriver.Edge() # Nesse caso, o caminho do WebDriver está salvo como variável do sistema.

  # Navegação até a página de temperatura
  driver.get("https://www.google.com.br/search?q=temperatura+s%C3%A3o+paulo")

  # Extração da temperatura e umidade
  elemento_temperatura = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[13]/div/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[1]/div[1]/div/div[1]/span[1]')
  elemento_umidade = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[13]/div/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[1]/div[2]/div[2]/span')

  # Conversão de elemento HTML para texto (caso necessário)
  temperatura = elemento_temperatura if isinstance(elemento_temperatura, str) else elemento_temperatura.text + '°C'
  umidade = elemento_umidade if isinstance(elemento_umidade, str) else elemento_umidade.text

  # Captura da data e hora atuais
  data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

  # Salvar os dados na planilha
  salvar_dados_na_planilha(data_hora, temperatura, umidade)

  # Mostra os dados salvos na label de mensagens
  label_mensagem.config(text=f"Dados salvos:\n{data_hora}\nTemperatura: {temperatura}\nUmidade: {umidade}")

  # Fechar o navegador
  driver.quit()

# - Configurações da janela
# Criação
janela = tk.Tk()

# Título da janela
janela.title("Weather Now")

# Tamanho da janela
janela.geometry("300x260")

# Impedir o redimensionamento
janela.resizable(False, True)

# - Widgets
# Imagem (Logo)
logo = Image.open("./static/images/logo.png")
logo = logo.resize((100, 100))
imagem_tk = ImageTk.PhotoImage(logo)

label_imagem = tk.Label(janela, image=imagem_tk)
label_imagem.pack()

# Texto
label = tk.Label(janela, text="Atualize a situação do clima em São Paulo!")
label.pack() # Adicionar o widget à janela

# Botão Salvar Clima
botao = tk.Button(janela, text="Buscar previsão", command=salvar_temperatura)
botao.pack(pady="8px")

# Mensagem
label_mensagem = tk.Label(janela, text="")
label_mensagem.pack(pady="5px")

# - Loop da aplicação para exibir a janela
janela.mainloop()
